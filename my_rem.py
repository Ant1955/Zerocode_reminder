import datetime
import logging
import random
import threading
import time
from datetime import datetime
from io import BytesIO

import pytz
import telebot
from openpyxl import load_workbook
from telegram import Update, KeyboardButton, ReplyKeyboardMarkup
from telegram.ext import CallbackContext
from timezonefinder import TimezoneFinder


def check_time_fmt(shdl_time):
    if len(shdl_time) != 5:
        return False
    shdl_hh = int(shdl_time[0:2])
    shdl_ddot = shdl_time[2]
    shdl_min = int(shdl_time[3:5])
    if 0 <= shdl_hh < 24 and shdl_ddot == ':' and 0 <= shdl_min < 60:
        return True
    else: 
        return False

topic = ["Пора на работу", "Duolingo пока едешь", "Что нового в почте", "Можешь попить кофе",
           "Сколько задач остались?", "Иногда люди обедают", "Начальник тебя видел?", "Составь план на завтра",
           "может уже домой?", "Зерокот что пишет?", "Домашку сделал?", "не стоит поздно ложиться спать"]

shedule = ["8:00", "8:45", "9:30", "11:30", "12:00", "15:00", "17:00", "23:38", "23:26", "22:54", "22:56", "23:36"]

shift = "99:"
for i in range(20):
    shedule.append(shift + str(i + 18))

# bot = telebot.TeleBot('7394207851:AAFzbujyT1ekzK2I2OOOD95XXL7hK0KON9M')
bot = telebot.TeleBot('7415862022:AAEvrljMTbCNODKkMjIPgP-CTnIue9fwahQ')

# Включаем логирование
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO)

# Создаем экземпляр TimezoneFinder
tf = TimezoneFinder()

@bot.message_handler(commands=['start'])
def start_message(message, update: Update, context: CallbackContext) -> None:
    keyboard = [[KeyboardButton("Send location", request_location=True)]]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True)
    update.message.reply_text('Please share your location to find out your local time.', reply_markup=reply_markup)
    bot.reply_to(message, "Привет! Я чат бот, который может напоминать учить уроки и делать домашку!\n"
                          "/shedule - задать новое расписание в чате\n"
                          "можно загрузить файл .xlsx где в первой колонке время, а во втрой тема\n"
                          "некоторое расписание для примера есть по умолчанию\n"
                          "/list - посмотреть текущее расписание\n"
                          "/fact - узнать случайную мысль (их не много)\n")
    global reminder_thread
    reminder_thread = threading.Thread(target=send_reminders, args=(message.chat.id,))
    reminder_thread.start()


# Обработчик местоположения
def location(update: Update, context: CallbackContext) -> None:
    user_location = update.message.location
    if user_location:
        lat = user_location.latitude
        lon = user_location.longitude

        # Определяем временную зону
        timezone_str = tf.timezone_at(lat=lat, lng=lon)
        if timezone_str:
            timezone = pytz.timezone(timezone_str)
            local_time = datetime.now(timezone)
            update.message.reply_text(f'Your local time is: {local_time.strftime("%Y-%m-%d %H:%M:%S")}')
        else:
            update.message.reply_text('Could not determine your timezone.')
    else:
        update.message.reply_text('Location not received.')

@bot.message_handler(commands=['fact'])
def action_request(message):
    list = ["Используй Duolingo - простой способ не забыть язык и тренировать память",
            "Реже смотри телевизор - лучше подпишись на ленту новостей по своему выбору",
            "Пиши на Python рассказывая идею программы chatGPT - в мире уже все написано",
            "Не очень доверяй Линуксу и программам от github - их писали такие же как ты"]
    random_fact = random.choice(list)
    bot.reply_to(message, f"Лови мысль: {random_fact}")

@bot.message_handler(commands=['shedule'])
def build_shedule(message):
    bot.reply_to(message, "построчно вводите время в формате НН:MM и тему напоминания через пробел,"
                          "обновленное расписание запустится тут же")
    mxtopic = len(topic)
    for i in range(mxtopic):
        dummy = shedule.pop(0)
    topic.clear()

@bot.message_handler(commands=['list'])
def list_shedule(message):
    chat_id = message.chat.id
    for i in range(len(topic)):
        bot.send_message(chat_id, f"{shedule[i]}\t {topic[i]}")

@bot.message_handler(content_types=['document'])
def handle_document(message):
    try:
        chat_id = message.chat.id
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        # Проверяем, что это файл Excel
        if message.document.mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or message.document.file_name.endswith('.xlsx'):
            mxtopic = len(topic)
            for i in range(mxtopic):
                dummy = shedule.pop(0)
            topic.clear()
            wb = load_workbook(BytesIO(downloaded_file))
            ws = wb.active  # Активный лист
            for row in ws.iter_rows(values_only=True):
                s_time = str(row[0])[:5]
                if check_time_fmt(s_time):
                    shedule.insert(len(topic), s_time)
                    topic.append(row[1])
        else:
            bot.reply_to(message, "Пожалуйста, отправьте файл в формате .xlsx.")
    except Exception as e:
        bot.reply_to(message, f"Произошла ошибка: {str(e)}")

@bot.message_handler(func=lambda message: True)
def handle_message(message):
    chat_id = message.chat.id
    text = message.text
    line = text.split()
    # надо прверять формат line[0]
    shdl_time = str(line[0])
    if check_time_fmt(shdl_time):
        shedule.insert(len(topic), shdl_time)
        topic.append(text[len(shdl_time):])
    else:
        bot.send_message(chat_id, "что-то не так c форматом времени, надо HH:MM\nпопробуйте снова")

def send_reminders(chat_id):
    while True:
        now = datetime.datetime.now().strftime('%H:%M')
        for k in range(len(topic)):
            if now == shedule[k]:
                markup = telebot.types.InlineKeyboardMarkup()
                like_button = telebot.types.InlineKeyboardButton("👍", callback_data=f"like_{k}")
                markup.add(like_button)
                bot.send_message(chat_id, topic[k], reply_markup=markup)
                time.sleep(61)
        time.sleep(3)

@bot.callback_query_handler(func=lambda call: call.data.startswith("like_"))
def callback_like(call):
    try:
        # Обработка лайка
        bot.answer_callback_query(call.id, "Спасибо за лайк!")
        bot.send_message(call.message.chat.id, f"Лайк добавлен {call.id}, спасибо! {call.data}")
    except Exception as e:
        bot.answer_callback_query(call.id, f"Произошла ошибка: {str(e)}")

bot.polling(non_stop=True)
